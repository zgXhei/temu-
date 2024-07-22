import json
from DrissionPage import ChromiumPage
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
import os
import cv2

# 自定义的正则匹配，根据需要爬取的图片进行修改，我的图片命名格式统一，所以爬取完十张进行一次换行
# 将按照格式命名好的图片直接上传到素材中心的全部素材就行，注意中间会出现访问人数过多或者漏图的错误，是拼多多服务器问题
patterns = {
    '主图0': re.compile(r'\b0\b'),  # 对提取到的图片名进行一个匹配方便后续存储
    '主图1': re.compile(r'\b1\b'),  # \b代表的是边界，后续不会和预览图中的数字进行冲突
    '尺寸图': re.compile(r'\b尺寸图\b'),
    '场景图2': re.compile(r'\b2\b'),
    '场景图3': re.compile(r'\b3\b'),
    '场景图4': re.compile(r'\b4\b'),
    '场景图5': re.compile(r'\b5\b'),

    '预览图24x32': re.compile(r'\b24x32\b'),  # 直接对图片名进行一个匹配,详细匹配方便尺寸的分类
    '预览图24x36': re.compile(r'\b24x36\b'),
    '预览图30x40': re.compile(r'\b30x40\b')
}

filename_sum = 0  # 文件名累加器
i = 6  # 定义轮播图插入的列
d = 0  # 对items的遍历进行一个统计
x = 7
size_sum1 = 7  # 定义预览图的插入行
size_sum2 = 8
size_sum3 = 9

folder_path = r"images"  # 图片存储地址
file_names = r"zip_img"  # 压缩图片地址

wb = load_workbook('商品上传模版.xlsx')  # 打开工作簿
ws = wb['模版']  # 打开工作表


def driver_chromium_page():
    driver = ChromiumPage()  # 实例化浏览器
    driver.get('https://seller.kuajingmaihuo.com/settle/site-main')  # 进主页
    driver.wait(2)  # 等待
    driver.ele('css:.site-main_btnGroup__3fEoG').click()  # 进后台
    driver.wait(4)

    # 使用try来对可能会出现的广告弹窗进行一个处理,如果有就进行一个点击,没有也能使程序继续运行
    try:
        # class="MDL_iconWrapper_5-109-0" 广告的x的class标签
        driver.ele('css:.MDL_iconWrapper_5-109-0', timeout=2).click()
        driver.wait(2)
    except Exception as e:
        # 如果没有就跳过不处理
        print('无广告')
    driver.ele('css:[data-report-click-text$=商品管理]').click()  # 进商品管理
    driver.wait(1)
    driver.ele('css:[data-report-click-text$=素材中心]').click()  # 进素材中心
    # driver.listen.start('marvel-mms/cn/api/kiana/gmp/bg/phoenix/api/material/page-query-material')  # 监听包
    # 对爬取数据多少条进行一个选择
    driver.ele('css:.index-module__pagination-wrapper___17qlp').ele(
        'css:.ST_headDropdownArrow_5-113-0').click()  # 点击, 这个位置的class属性可能不同的人不一样,按照需求修改
    driver.wait(1)
    driver.listen.start('marvel-mms/cn/api/kiana/gmp/bg/phoenix/api/material/page-query-material')  # 监听包
    driver.ele('@text()=200').click()  # 对需要爬取的数据进行选择，500,200,100,40,20,10
    driver.wait(3)
    resp = driver.listen.wait()  # 等待数据包加载完毕
    text1 = resp.response.body  # 获取数据包的响应数据
    return text1


def process_json(text2):
    re_text = re.findall('(.*)', str(text2))[0].replace("'", '"')  # 对得到数据进行处理，去除单引号
    re_text = re_text.replace('None', 'null')  # 替换None
    re_text = re_text.replace('False', 'false')  # 替换False
    re_text = re_text.replace('True', 'true')  # 替换True
    json_data = json.loads(re_text)  # 处理成json格式方便后续提取数据
    items1 = json_data['result']['materialList']  # 获取到json_data中['result']['materialList']中的数据里面存放着id和图片名
    return items1


if __name__ == '__main__':
    text = driver_chromium_page()
    items = process_json(text)

    for item in items:
        # 遍历去获取他的图片名和需要的图片id
        for type_name, pattern in patterns.items():
            # 获取到我定义的正则字典中的key和value
            match = pattern.search(item['materialName'])
            # 用value去匹配数据中的图片名
            if match:
                # 如果找到匹配项，则执行相应的操作
                if type_name == '主图0':
                    img_url = item['imgUrl']  # 获取主图的链接
                    url_name = str(item['id'])  # 将主图id作为图片的文件名
                    # 对图片进行一个持久化的操作，将他储存在images这个文件夹中
                    with open(f'images/{url_name}.jpg', 'wb') as f:
                        f.write(requests.get(img_url).content)
                    img_path = os.path.join(folder_path, url_name + '.jpg')  # 拼接出刚刚保存的图片地址

                    img = cv2.imread(img_path)  # 对图片进行一个读取并存入变量img中
                    resized = cv2.resize(img, (140, 140))  # 使用cv2对图片进行一个压缩，重新定义图片尺寸为150,150
                    zip_img_path = os.path.join(file_names, url_name + '.jpg')  # 拼接出存入图片地址
                    cv2.imwrite(zip_img_path, resized)  # 将刚刚压缩的图片存入压缩图片地址中

                    ws.add_image(img=Image(zip_img_path), anchor=f'C{x}')  # 在进行一个图片插入Excel中，将图片插入C7中

                    ws[f'BP{i}'] = str(item['id'])  # id插入到轮播图插入列中
                if type_name == '主图1':
                    ws[f'BQ{i}'] = str(item['id'])
                if type_name == '尺寸图':
                    ws[f'BR{i}'] = str(item['id'])
                if type_name == '场景图2':
                    ws[f'BS{i}'] = str(item['id'])
                if type_name == '场景图3':
                    ws[f'BT{i}'] = str(item['id'])
                if type_name == '场景图4':
                    ws[f'BU{i}'] = str(item['id'])
                if type_name == '预览图30x20':
                    ws[f'BV{i}'] = str(item['id'])

                if type_name == '预览图24x32':
                    ws[f'AN{size_sum1}'] = '24x32'
                    ws[f'AO{size_sum1}'] = str(item['id'])  # 没有成功插入请检查你的正则和图片名称是否正确可能是乘号也可能是小写x
                if type_name == '预览图24x36':
                    ws[f'AN{size_sum2}'] = '24x36'
                    ws[f'AO{size_sum2}'] = str(item['id'])
                if type_name == '预览图30x40':
                    ws[f'AN{size_sum3}'] = '40x30'
                    ws[f'AO{size_sum3}'] = str(item['id'])

        d += 1  # 对item跟踪器进行累加，如果达到被10整除时,一次保存十条id
        if d % 10 == 0:
            filename_sum += 1  # 文件名累加器进行累加
            i += 4
            c = 7
            x += 4
            size_sum1 += 4
            size_sum2 += 4
            size_sum3 += 4
            # size_sum1, size_sum2, size_sum3 = size_sum1+4, size_sum2+4, size_sum3+4
            print(f'完成爬取链接第{filename_sum}套')  # 对运行结果进行一次可视化输出
    wb.save(f'商品上传模版({filename_sum}).xlsx')  # 保存格式为‘竖模({filename_sum}).xlsx’
